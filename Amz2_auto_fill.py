from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import pandas as pd

Sales_details = pd.read_excel(r"E:\python\项目\李青测评\订单.xlsx")
Comparison_table = pd.read_excel(r"E:\python\项目\李青测评\对照表.xlsx")#读取对照表
Sales_details['订单币种'].replace('GBP', 9,inplace = True)#币种变更为汇率
Sales_details['订单币种'].replace('CAD', 5,inplace = True)#币种变更为汇率
Sales_details['订单币种'].replace('EUR', 7.7,inplace = True)#币种变更为汇率
Sales_details['订单币种'].replace('JPY', 0.058,inplace = True)#币种变更为汇率
Sales_details['订单币种'].replace('USD', 6.3,inplace = True)#币种变更为汇率
Sales_details['订购日期'] = Sales_details['订购日期'].astype("str").str[0:10] #截取年-月，变更日期格式，方便透视
#报表合并
merge_Sales_details = pd.merge(
                                Sales_details,Comparison_table,
                                how='left',
                                left_on=('ASIN','MSKU'),
                                right_on=('ASIN','平台SKU'))#合并对照表
merge_Sales_details.to_excel(r"E:\python\项目\李青测评\合并表.xlsx")

#合并表中取值
wb = load_workbook (r'E:\python\项目\李青测评\合并表.xlsx')
ws = wb['Sheet1']
sheet = wb.active
a = sheet.max_row
Time = []   #时间
Shop_name = []#店铺名
Sku = []  #SKU
Product_name = [] #产品名
Currency = [] #币种
Unit_Price = [] #单价
Asin = []  #ASIN
Order_number = [] #订单号
Business = [] #平台
Salesman = [] #销售人员
Department = [] #部门
Director =[] #负责人
Exchange_rate = [] #汇率
i=0
while i < a:
    i = i+1
    Time.insert(i, ws.cell (i,9).value)
    Shop_name.insert(i, ws.cell(i, 2).value)
    Sku.insert(i, ws.cell(i, 22).value)
    Product_name.insert(i, ws.cell(i, 23).value)
    Currency.insert(i, ws.cell(i, 4).value)
    Unit_Price.insert(i, ws.cell(i, 26).value)
    Asin.insert(i, ws.cell(i, 19).value)
    Order_number.insert(i, ws.cell(i, 5).value)
    Business.insert(i, ws.cell(i, 53).value)
    Salesman.insert(i, ws.cell(i, 50).value)
    Department.insert(i, ws.cell(i, 54).value)
    Director.insert(i, ws.cell(i, 55).value)
    Exchange_rate.insert(i, ws.cell(i, 4).value)
#追加合并表中数据到新表中去
wbtest = load_workbook (r'E:\python\项目\李青测评\模板.xlsx')
ws = wbtest['Sheet1']
sheet = wb.active
i = 1
while i < a:
    ws.cell (i+1,2).value = Time[i]
    ws.cell(i + 1, 6).value = Shop_name[i]
    ws.cell(i + 1, 7).value = Sku[i]
    ws.cell(i + 1, 9).value = Product_name[i]
    ws.cell(i + 1, 13).value = Currency[i]
    ws.cell(i + 1, 11).value = Unit_Price[i]
    ws.cell(i + 1, 19).value = Asin[i]
    ws.cell(i + 1, 20).value = Order_number[i]
    ws.cell(i + 1, 1).value = Business[i]
    ws.cell(i + 1, 3).value = Salesman[i]
    ws.cell(i + 1, 4).value = Department[i]
    ws.cell(i + 1, 5).value = Director[i]
    ws.cell(i + 1, 12).value = Exchange_rate[i]
    i = i+1
wbtest.save (r'E:\python\项目\李青测评\模板.xlsx')
